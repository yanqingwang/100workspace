# -*- coding: utf-8 -*-
#
from openpyxl import load_workbook
import time
import csv

from os import chdir,listdir


class HandleFiles():
    def __init__(self):
        self.path = 'C:/temp/xlsx/'
        self.prefix = 'ori_'
        self.out_prefix = 'Output_res_'
        self.target_cells = [{'RU':'A8'},{'Country':'F8'}]
        self.output_data = []

    def read_files(self,lv_file):
        list = {}
        try:
            # 1.打开 Excel 表格并获取表格名称
            workbook = load_workbook(filename=lv_file)
            # print(workbook.sheetnames)
            # 2.通过 sheet 名称获取表格
            sheet = workbook["Sheet1"]
            # print(sheet)
            # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
            # print('Dimensions',sheet.dimensions)

            for target1 in self.target_cells:
                for key in target1.keys():
                    list[key] = sheet[target1.get(key)].value
                    # print(list)

            self.output_data.append(list)
            # print(self.output)

        except Exception as e:
            print('Exception:', e)

    def main(self):
        # 读取文件
        filenames = listdir(self.path)
        for filename in filenames:
            print(filename)
            if filename.startswith('ori_'):
                file_withpath =self.path  + filename
                self.read_files(file_withpath)

        outfile = self.path + self.out_prefix + '.csv'
        with open(outfile,"w") as f:
            # w = csv.writer(f,delimiter="\n")
            # w.writerow(self.output_data)
            w = csv.writer(f,delimiter="|")
            fieldnames = self.output_data[0].keys()  # solve the problem to automatically write the header
            w.writerow(fieldnames)
            for row in self.output_data:
                # w.writerow(row.keys())
                w.writerow(row.values())


if __name__ == '__main__':
    time1 = time.time()
    file_merge = HandleFiles()
    file_merge.main()
    print("Done, Total running time", time.time() - time1)

