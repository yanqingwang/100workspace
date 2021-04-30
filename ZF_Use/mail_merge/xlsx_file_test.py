# -*- coding: utf-8 -*-
#
from openpyxl import load_workbook
import time


class handle_files():
    def __init__(self):
        self.path = 'C:/temp/xlsx/'
        self.prefix = 'ori_'
        self.out_prefix = 'Output_res_'
        self.target_cells = [{},{}]
        self.output = []

    def read_files(self):
        try:
            # 1.打开 Excel 表格并获取表格名称
            workbook = load_workbook(filename=self.path + "ori_Global_DIV_RU.xlsx")
            print(workbook.sheetnames)
            # 2.通过 sheet 名称获取表格
            sheet = workbook["Sheet1"]
            print(sheet)
            # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
            print('Dimensions',sheet.dimensions)
        # 4.获取表格内某个格子的数据
            # 1 sheet["A1"]方式
            cell1 = sheet["A1"]
            cell2 = sheet["C11"]
            print('Cell value',cell1.value, cell2.value)
            # 4.2sheet.cell(row=, column=)方式
            cell1 = sheet.cell(row = 1,column = 1)
            cell2 = sheet.cell(row = 11,column = 3)
            print('Cell value2',cell1.value, cell2.value)

            # 5. 获取一系列格子
            # 获取 A1:C2 区域的值
            cell = sheet["A1:C2"]
            print(cell)
            for i in cell:
               for j in i:
                   print(j.value)

        except Exception as e:
            print('Exception:', e)

if __name__ == '__main__':
    time1 = time.time()
    file_merge = handle_files()
    file_merge.read_files()
    print("Done, Total running time", time.time() - time1)

